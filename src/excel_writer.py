"""Denetlenebilir Excel çıktısı üretir: 4 sekme.

  1. Ham Bulgular              — her bir terim eşleşmesi (sayfa no + alıntı ile)
  2. Şirket-Yıl-RaporTürü Özeti — 100 şirket × 3 yıl × 2 rapor türü TAM matris
  3. Kategori 1-15 Matrisi      — hangi Kapsam 3 kategorisi hangi rapor için bulundu
  4. Eksik Raporlar             — rapor bulunamayan/indirilemeyen şirket-yıl-tür'ler

Her satır kaynağına (sayfa numarası, alıntı, kaynak URL) kadar izlenebilir
olacak şekilde tasarlanmıştır.
"""
from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

import config
from src.maturity_classifier import tag_finding
from src.models import Company, Finding, MaturityResult, OlgunlukEtiketi, ReportRef, ReportStatus

_HEADER_FILL = PatternFill(start_color="FF1F4E78", end_color="FF1F4E78", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFFFF", bold=True)
_WRAP = Alignment(wrap_text=True, vertical="top")
_MAX_CATEGORY = 15


def _write_header(ws: Worksheet, headers: list[str], widths: list[int]) -> None:
    for col_idx, (name, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _company_unvan(companies_by_kod: dict[str, Company], kod: str) -> str:
    c = companies_by_kod.get(kod)
    return c.unvan if c else "(BIST100 dışı / bilinmiyor)"


# --------------------------------------------------------------------------
# Sekme 1: Ham Bulgular
# --------------------------------------------------------------------------

def _write_ham_bulgular(
    ws: Worksheet,
    companies_by_kod: dict[str, Company],
    findings: list[Finding],
    manifest_by_key: dict[tuple[str, int, str], ReportRef],
) -> None:
    headers = [
        "Şirket Kodu", "Şirket Ünvanı", "Yıl", "Rapor Türü", "Sayfa No",
        "Eşleşen Terim", "Bağlam Alıntısı", "Olgunluk Sinyali", "Kategori No",
    ]
    widths = [12, 45, 8, 20, 10, 22, 70, 24, 12]
    _write_header(ws, headers, widths)

    row_idx = 2
    for f in sorted(findings, key=lambda f: (f.kod, f.yil, f.rapor_turu, f.sayfa_no)):
        ws.cell(row=row_idx, column=1, value=f.kod)
        ws.cell(row=row_idx, column=2, value=_company_unvan(companies_by_kod, f.kod))
        ws.cell(row=row_idx, column=3, value=f.yil)
        ws.cell(row=row_idx, column=4, value=f.rapor_turu)
        ws.cell(row=row_idx, column=5, value=f.sayfa_no)
        ws.cell(row=row_idx, column=6, value=f.eslesen_terim)
        alinti_cell = ws.cell(row=row_idx, column=7, value=f.alinti)
        alinti_cell.alignment = _WRAP
        ws.cell(row=row_idx, column=8, value=tag_finding(f).value)
        ws.cell(row=row_idx, column=9, value=f.kategori_no)
        row_idx += 1


# --------------------------------------------------------------------------
# Ortak: tam matris iterasyonu (100 şirket × yıl × rapor türü)
# --------------------------------------------------------------------------

def _full_matrix(
    companies: list[Company],
    manifest_by_key: dict[tuple[str, int, str], ReportRef],
    maturity_by_key: dict[tuple[str, int, str], MaturityResult],
):
    for company in companies:
        for yil in config.SCAN_YEARS:
            for rapor_turu in config.REPORT_TYPES:
                key = (company.kod, yil, rapor_turu)
                ref = manifest_by_key.get(key)
                maturity = maturity_by_key.get(key)
                yield company, yil, rapor_turu, ref, maturity


# --------------------------------------------------------------------------
# Sekme 2: Şirket-Yıl-RaporTürü Özeti
# --------------------------------------------------------------------------

def _write_ozet(
    ws: Worksheet,
    companies: list[Company],
    manifest_by_key: dict[tuple[str, int, str], ReportRef],
    maturity_by_key: dict[tuple[str, int, str], MaturityResult],
) -> None:
    headers = [
        "Şirket Kodu", "Şirket Ünvanı", "Yıl", "Rapor Türü", "Rapor Durumu",
        "Olgunluk Etiketi", "Kategori Kapsamı (1-15)", "Gerekçe",
        "Kaynak Türü", "Kaynak URL", "Manuel Teyit Notu",
    ]
    widths = [12, 45, 8, 20, 14, 24, 22, 45, 16, 45, 30]
    _write_header(ws, headers, widths)

    row_idx = 2
    for company, yil, rapor_turu, ref, maturity in _full_matrix(companies, manifest_by_key, maturity_by_key):
        durum = ref.durum.value if ref else ReportStatus.BULUNMADI.value
        etiket = maturity.etiket.value if maturity else (OlgunlukEtiketi.ACIKLAMA_YOK.value if durum == ReportStatus.INDIRILDI.value else "-")
        kategori_str = ", ".join(f"Kategori {n}" for n in maturity.kategori_kapsami) if maturity else ""

        ws.cell(row=row_idx, column=1, value=company.kod)
        ws.cell(row=row_idx, column=2, value=company.unvan)
        ws.cell(row=row_idx, column=3, value=yil)
        ws.cell(row=row_idx, column=4, value=rapor_turu)
        ws.cell(row=row_idx, column=5, value=durum)
        ws.cell(row=row_idx, column=6, value=etiket)
        ws.cell(row=row_idx, column=7, value=kategori_str)
        gerekce_cell = ws.cell(row=row_idx, column=8, value=maturity.gerekce if maturity else "")
        gerekce_cell.alignment = _WRAP
        ws.cell(row=row_idx, column=9, value=ref.kaynak_turu if ref else "")
        if ref and ref.kaynak_url:
            url_cell = ws.cell(row=row_idx, column=10, value=ref.kaynak_url)
            url_cell.hyperlink = ref.kaynak_url
            url_cell.font = Font(color="FF0563C1", underline="single")
        note = maturity.manuel_teyit_notu if maturity else (config.MATURITY_REVIEW_NOTE if durum == ReportStatus.INDIRILDI.value else "")
        ws.cell(row=row_idx, column=11, value=note)
        row_idx += 1


# --------------------------------------------------------------------------
# Sekme 3: Kategori 1-15 Matrisi
# --------------------------------------------------------------------------

def _write_kategori_matrisi(
    ws: Worksheet,
    companies: list[Company],
    manifest_by_key: dict[tuple[str, int, str], ReportRef],
    maturity_by_key: dict[tuple[str, int, str], MaturityResult],
) -> None:
    base_headers = ["Şirket Kodu", "Şirket Ünvanı", "Yıl", "Rapor Türü", "Rapor Durumu"]
    kategori_headers = [f"Kategori {n}" for n in range(1, _MAX_CATEGORY + 1)]
    headers = base_headers + kategori_headers
    widths = [12, 45, 8, 20, 14] + [10] * _MAX_CATEGORY
    _write_header(ws, headers, widths)

    row_idx = 2
    for company, yil, rapor_turu, ref, maturity in _full_matrix(companies, manifest_by_key, maturity_by_key):
        durum = ref.durum.value if ref else ReportStatus.BULUNMADI.value
        ws.cell(row=row_idx, column=1, value=company.kod)
        ws.cell(row=row_idx, column=2, value=company.unvan)
        ws.cell(row=row_idx, column=3, value=yil)
        ws.cell(row=row_idx, column=4, value=rapor_turu)
        ws.cell(row=row_idx, column=5, value=durum)

        kapsam = set(maturity.kategori_kapsami) if maturity else set()
        for offset, n in enumerate(range(1, _MAX_CATEGORY + 1)):
            ws.cell(row=row_idx, column=6 + offset, value="✓" if n in kapsam else "")
        row_idx += 1


# --------------------------------------------------------------------------
# Sekme 4: Eksik Raporlar
# --------------------------------------------------------------------------

def _write_eksik_raporlar(
    ws: Worksheet,
    companies: list[Company],
    manifest_by_key: dict[tuple[str, int, str], ReportRef],
) -> None:
    headers = ["Şirket Kodu", "Şirket Ünvanı", "Yıl", "Rapor Türü", "Durum", "Hata Mesajı"]
    widths = [12, 45, 8, 20, 14, 60]
    _write_header(ws, headers, widths)

    row_idx = 2
    for company in companies:
        for yil in config.SCAN_YEARS:
            for rapor_turu in config.REPORT_TYPES:
                ref = manifest_by_key.get((company.kod, yil, rapor_turu))
                durum = ref.durum if ref else ReportStatus.BULUNMADI
                if durum not in (ReportStatus.BULUNMADI, ReportStatus.HATA):
                    continue
                ws.cell(row=row_idx, column=1, value=company.kod)
                ws.cell(row=row_idx, column=2, value=company.unvan)
                ws.cell(row=row_idx, column=3, value=yil)
                ws.cell(row=row_idx, column=4, value=rapor_turu)
                ws.cell(row=row_idx, column=5, value=durum.value)
                ws.cell(row=row_idx, column=6, value=(ref.hata_mesaji if ref else None) or "")
                row_idx += 1


# --------------------------------------------------------------------------
# Genel giriş noktası
# --------------------------------------------------------------------------

def write_workbook(
    output_path,
    companies: list[Company],
    manifest: list[ReportRef],
    findings: list[Finding],
    maturity_results: list[MaturityResult],
) -> None:
    companies_by_kod = {c.kod: c for c in companies}
    manifest_by_key = {r.key: r for r in manifest}
    maturity_by_key = {(m.kod, m.yil, m.rapor_turu): m for m in maturity_results}

    wb = Workbook()
    wb.remove(wb.active)

    _write_ham_bulgular(wb.create_sheet("Ham Bulgular"), companies_by_kod, findings, manifest_by_key)
    _write_ozet(wb.create_sheet("Şirket-Yıl-RaporTürü Özeti"), companies, manifest_by_key, maturity_by_key)
    _write_kategori_matrisi(wb.create_sheet("Kategori 1-15 Matrisi"), companies, manifest_by_key, maturity_by_key)
    _write_eksik_raporlar(wb.create_sheet("Eksik Raporlar"), companies, manifest_by_key)

    output_path = config.OUTPUT_DIR / output_path if not str(output_path).startswith("/") else output_path
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)


def main() -> None:
    from src.downloader import load_manifest
    from src.inventory import load_bist100
    from src.maturity_classifier import classify_all
    from src.text_miner import load_findings

    companies = load_bist100()
    manifest = load_manifest()
    findings = load_findings()
    maturity_results = classify_all(findings)

    write_workbook(config.OUTPUT_XLSX, companies, manifest, findings, maturity_results)
    print(f"[excel_writer] Yazıldı: {config.OUTPUT_XLSX}")


if __name__ == "__main__":
    main()

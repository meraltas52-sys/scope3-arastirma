"""Her şirket-yıl-rapor türü için PDF indirir. İdempotent: zaten indirilmiş
ve geçerli bir dosya varsa tekrar indirmez.

Rapor referansları iki kaynaktan derlenir (KGK önceliklidir - bkz.
build_ir_page_refs'in existing_keys ile KGK'da zaten çözülmüş kombinasyonları
atlaması):
  1. KGK Rapor Envanteri (config.TSRS_ENVANTERI_XLSX) — önceden derlenmiş,
     "Önerilen BIST Kodu" ile şirket eşleştirmesi yapılmış 2023-2025
     sürdürülebilirlik/entegre rapor bağlantıları. NOT: kgk.gov.tr bu ortamdan
     gelen isteklerin çoğuna 500 döndürüyor; kaynağı KAPATMAK yerine (gerçek
     TSRS raporlarını kaçırır) hızlı-başarısızlık politikası uygulanır -
     KGK_MAX_RETRIES/KGK_TIMEOUT_SECONDS (bkz. config.py), config.ENABLE_KGK_SOURCE
     ile tamamen de kapatılabilir.
  2. Şirketin kendi yatırımcı ilişkileri sayfası (inventory.py cache'i) —
     ana sayfadan başlayıp rapor/yatırımcı ilişkileri ile alakalı görünen
     bağlantıları izleyen çok sayfalı bir tarama ile (bkz.
     discover_pdf_links_from_ir_page) PDF bağlantıları aranır. Sadece KGK'da
     bulunamayan kombinasyonlar için denenir.

Her iki kaynak için de bulunamayan şirket-yıl-rapor türü kombinasyonu
"Rapor Yok" olarak manifest'e yazılır (excel_writer.py "Eksik Raporlar"
sekmesinde kullanır) — atlanmaz, işaretlenir. Başarıyla indirilen her PDF için
ayrıca TSRS uygunluk kontrolü yapılır (bkz. check_tsrs_compliance) - IR sitesi
kaynaklı indirmeler KGK'nın aksine başlık ön-doğrulamasından geçmediği için bu
kontrol, "gerçekten TSRS raporu mu yoksa alakasız bir belge mi" ayrımını sağlar.

NOT (ağ erişimi): Bu modüldeki tüm indirme/tarama işlemleri canlı HTTP
istekleri gerektirir. Dış ağa kapalı bir ortamda çalıştırıldığında her
istek zaman aşımına uğrar/başarısız olur; her başarısızlık o rapor için
"Hata" ya da "Rapor Yok" olarak işaretlenir, akış durmaz.
"""
from __future__ import annotations

import logging
import re
import time
from dataclasses import asdict, replace
from pathlib import Path
from urllib.parse import urljoin, urlparse

import openpyxl
import requests

import config
from src.inventory import load_cache as load_ir_cache
from src.models import Company, InvestorRelationsInfo, ReportRef, ReportStatus
from src.progress import ProgressStore, load_json, run_in_chunks, save_json

logger = logging.getLogger(__name__)

try:
    from bs4 import BeautifulSoup
except ImportError:  # pragma: no cover - requirements.txt'de var, savunma amaçlı
    BeautifulSoup = None

try:
    import pymupdf as fitz  # PyMuPDF (eski adı `fitz`) - TSRS uygunluk kontrolü için
except ImportError:  # pragma: no cover - requirements.txt'de var, savunma amaçlı
    fitz = None


# Tarayıcı gibi görünen istek başlıkları — bazı sunucular (örn. KGK) sade
# "User-Agent" içeren istekleri bot olarak algılayıp 500/403 dönebiliyor.
BROWSER_HEADERS = {
    "User-Agent": config.HTTP_USER_AGENT,
    "Accept": "application/pdf,application/octet-stream,text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "tr-TR,tr;q=0.9,en-US;q=0.8,en;q=0.7",
    "Referer": "https://www.kgk.gov.tr/",
    "Connection": "keep-alive",
}

# İstekler arası minimum bekleme (saniye) — hızlı art arda istek atmak
# bazı sunucularda geçici engellemeye/500 hatasına yol açabiliyor.
REQUEST_DELAY_SECONDS = 1.5


# --------------------------------------------------------------------------
# Rapor türü sınıflandırma yardımcıları
# --------------------------------------------------------------------------

def infer_report_type(title: str) -> str:
    """Rapor başlığından 3 kategoriden birini seçer (bkz. config.REPORT_TYPES).

    KGK listesi "Entegre Faaliyet Raporu" gibi birleşik başlıklar da içerir;
    bunlar sürdürülebilirlik içeriği taşıdığından sürdürülebilirlik kovalarından
    (TSRS/Gönüllü) birine dahil edilir. Yalnızca "sürdürülebilirlik/entegre"
    geçmeyen saf "faaliyet raporu" başlıkları "Faaliyet Raporu" olarak sınıflanır.

    NOT: Başlıkta "TSRS" geçmese bile PDF'in ilk 3 sayfası TSRS standardına atıf
    yapıyorsa rapor yine de TSRS Uyumlu kategorisine YÜKSELTİLİR - ama bu ancak
    dosya indirildikten sonra mümkündür (bkz. downloader.run()'daki
    _promote_if_tsrs_content). Bu fonksiyon yalnızca başlık bazlı ilk tahmini yapar.
    """
    t = title.lower()
    if "tsrs" in t:
        return config.REPORT_TYPE_TSRS
    if "sürdürülebilirlik" in t or "entegre" in t or "surdurulebilirlik" in t:
        return config.REPORT_TYPE_GONULLU
    if "faaliyet" in t:
        return config.REPORT_TYPE_FAALIYET
    return config.REPORT_TYPE_GONULLU


# --------------------------------------------------------------------------
# Kaynak 1: KGK Rapor Envanteri
# --------------------------------------------------------------------------

_UNVAN_STOPWORDS = {"A.Ş", "A.O", "VE"}
# Python'un yerelden bağımsız str.upper() metodu Türkçe "i" harfini ASCII "I"
# (noktasız) yapar, doğrusu ise "İ" (noktalı büyük I) olmalı - aksi halde
# "Mavi" -> "MAVI" olur ama ünvan verisindeki "MAVİ" ile hiç eşleşmez.
_TR_UPPER_MAP = str.maketrans({"i": "İ", "ı": "I"})


def _tr_upper(text: str) -> str:
    """Eşleştirme amaçlı büyük harfe çevirir; noktalı/noktasız I ayrımını
    yok sayar (örn. "QUA Granite" ünvanındaki İngilizce "Granite" sözcüğü
    kaynak veride noktasız "I" ile yazılmışken rapor başlığında Türkçe
    kurala göre "GRANİTE" üretilebilir - ikisini de aynı kelime say)."""
    return text.translate(_TR_UPPER_MAP).upper().replace("İ", "I")


def _unvan_words(unvan: str) -> set[str]:
    # >=3 karakter: "MLP" (MLP Sağlık Hizmetleri -> "MLP Care"), "TAV" gibi
    # kısa marka kısaltmaları da ayırt edici olabiliyor; 1-2 harfli ekleri
    # (VE, A.Ş gibi) zaten _UNVAN_STOPWORDS eler.
    return {w for w in re.split(r"[^A-ZÇĞİÖŞÜ0-9]+", _tr_upper(unvan)) if len(w) >= 3 and w not in _UNVAN_STOPWORDS}


def _title_words(title: str) -> set[str]:
    return {w for w in re.split(r"[^A-ZÇĞİÖŞÜ0-9]+", _tr_upper(title)) if w}


def _has_shared_word(company_words: set[str], baslik_words: set[str], min_prefix: int = 5) -> bool:
    """Ünvan kelimesi ile başlık kelimesi eşleşiyor mu (tam veya marka kısaltması).

    Şirketler resmi ünvanlarından kısaltılmış marka adlarıyla anılabilir
    (örn. "TÜRK TELEKOMÜNİKASYON A.Ş." -> rapor başlığında "Türk Telekom").
    Bu yüzden tam eşleşmenin yanı sıra, en az `min_prefix` karakter paylaşan
    önek eşleşmesi de kabul edilir.
    """
    for cw in company_words:
        for bw in baslik_words:
            if cw == bw:
                return True
            shorter, longer = (cw, bw) if len(cw) <= len(bw) else (bw, cw)
            if len(shorter) >= min_prefix and longer.startswith(shorter):
                return True
    return False


_GENERIC_WORD_MIN_COUNT = 5


def _build_generic_unvan_words(companies_by_kod: dict[str, Company]) -> set[str]:
    """Örneklemde çok sayıda şirketin ünvanında geçen sektör/hukuki eklerini döner.

    "SANAYİ", "TİCARET", "HOLDİNG", "TÜRK", "TÜRKİYE", "BANKASI" gibi kelimeler
    onlarca şirket ünvanında ortak geçtiği için tek başına ayırt edici bir
    eşleşme sinyali değildir (örn. "TÜRK" hem TTKOM'un hem Kuveyt Türk'ün
    ünvanında var). Eşik 5 olarak seçildi: BIST100 örnekleminde bu eşiğin
    altındaki tekrarlar ("ANADOLU" 4 şirkette geçse de "Anadolu Sigorta" gibi
    gerçekten ayırt edici bir marka sözcüğü olabiliyor) gerçek eşleşmeleri
    yanlışlıkla eleyip false-positive üretiyordu; eşiğin üstündekiler ise
    tutarlı biçimde salt sektör/hukuki dolgu kelimesi. Sabit bir engelleme
    listesi yerine örneklemden dinamik hesaplanır, böylece bakım gerektirmez.
    """
    counts: dict[str, int] = {}
    for company in companies_by_kod.values():
        for w in _unvan_words(company.unvan):
            counts[w] = counts.get(w, 0) + 1
    return {w for w, n in counts.items() if n >= _GENERIC_WORD_MIN_COUNT}


def load_tsrs_envanteri_refs(companies_by_kod: dict[str, Company], path=config.TSRS_ENVANTERI_XLSX) -> list[ReportRef]:
    """Önceden derlenmiş KGK envanterini BIST100 örneklemiyle eşleştirip ReportRef üretir.

    BIST100 dışındaki eşleşmeler SESSİZCE ATLANMAZ: sadece bu fonksiyonun
    döndürdüğü listeye girmezler (örneklem dışı oldukları için); tam envanter
    karşılaştırması için bkz. `diff_tsrs_with_bist100`.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Rapor Envanteri"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}
    generic_words = _build_generic_unvan_words(companies_by_kod)

    refs: list[ReportRef] = []
    for row in rows[1:]:
        yil = row[idx["Yıl"]]
        baslik = row[idx["Rapor Başlığı"]] or ""
        url = row[idx["Kaynak URL"]]
        onerilen_kod = (row[idx.get("Önerilen BIST Kodu", -1)] or "").strip().upper() if idx.get("Önerilen BIST Kodu") is not None else ""
        if not yil or not url or onerilen_kod not in companies_by_kod:
            continue
        # Sanity check: rapor başlığı ile şirket ünvanı arasında en az bir
        # ayırt edici (örneklemde tekil geçen) ortak kelime olmalı - yoksa
        # "Önerilen BIST Kodu" sütunundaki otomatik eşleştirme hatalı olabilir
        # (örn. TTKOM -> Kuveyt Türk gibi; "TÜRK" iki ünvanda da geçtiği için
        # generic_words'te elenir ve gerçek çakışma tespit edilir).
        company_unvan = companies_by_kod[onerilen_kod].unvan
        baslik_words = _title_words(str(baslik))
        company_words = _unvan_words(company_unvan) - generic_words
        if company_words and not _has_shared_word(company_words, baslik_words):
            logger.warning(
                "Şüpheli eşleşme atlandı: %s koduna önerilen rapor '%s' başlıkla uyuşmuyor.",
                onerilen_kod, baslik,
            )
            continue
        refs.append(
            ReportRef(
                kod=onerilen_kod,
                yil=int(yil),
                rapor_turu=infer_report_type(baslik),
                kaynak_url=url,
                kaynak_turu="KGK",
                durum=ReportStatus.INDIRILECEK,
            )
        )
    return refs


def diff_tsrs_with_bist100(companies_by_kod: dict[str, Company], path=config.TSRS_ENVANTERI_XLSX) -> dict:
    """TSRS envanterindeki 150 raporu BIST100 örneklemiyle karşılaştırır.

    Döner: {
      "bist100_disi_satirlar": [...],       # envanterde var ama BIST100'de yok
      "bist100_raporu_olmayan_sirketler": [...],  # BIST100'de var, envanterde HİÇ raporu yok
    }
    Envanteri SİLMEZ/değiştirmez — sadece raporlar.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Rapor Envanteri"]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    bist100_disi = []
    kodlar_envanterde = set()
    for row in rows[1:]:
        onerilen_kod = (row[idx.get("Önerilen BIST Kodu", -1)] or "").strip().upper()
        baslik = row[idx["Rapor Başlığı"]]
        if onerilen_kod and onerilen_kod not in companies_by_kod:
            bist100_disi.append({"kod": onerilen_kod, "baslik": baslik})
        elif onerilen_kod:
            kodlar_envanterde.add(onerilen_kod)

    hic_raporu_olmayan = sorted(set(companies_by_kod) - kodlar_envanterde)
    return {
        "bist100_disi_satirlar": bist100_disi,
        "bist100_raporu_olmayan_sirketler": hic_raporu_olmayan,
    }


# --------------------------------------------------------------------------
# Kaynak 2: Şirket yatırımcı ilişkileri sayfası taraması
# --------------------------------------------------------------------------

_PDF_LINK_RE = re.compile(r'href=["\']([^"\']+\.pdf)["\']', re.IGNORECASE)

# Ana sayfanın kendisinde neredeyse hiç zaman doğrudan .pdf bağlantısı olmaz;
# gerçek sürdürülebilirlik/faaliyet raporları genelde "Yatırımcı İlişkileri
# > Raporlar" gibi birkaç tık derinde. Bu yüzden sadece verilen sayfayı değil,
# bu anahtar kelimelerle eşleşen bağlantıları da (aynı alan adında) izleyen
# en-iyi-önce (best-first) bir mini tarayıcı kullanılır.
_HUB_LINK_KEYWORDS = (
    "yatirimci", "investor", "surdurulebilirlik", "sustainability", "esg",
    "faaliyet", "annual", "rapor", "report", "finansal", "financial", "kurumsal",
)

_ASCII_FOLD_MAP = str.maketrans({
    "ı": "i", "İ": "i", "ş": "s", "Ş": "s", "ğ": "g", "Ğ": "g",
    "ü": "u", "Ü": "u", "ö": "o", "Ö": "o", "ç": "c", "Ç": "c",
})


def _fold_ascii(text: str) -> str:
    return text.translate(_ASCII_FOLD_MAP).lower()


_TR_SECOND_LEVEL_TLDS = {"com", "org", "gov", "net", "edu", "k12", "gen", "web", "info", "bel", "tv", "biz"}


def _registrable_domain(netloc: str) -> str:
    """Bir netloc'un kayıtlı alan adı kısmını kabaca döner (alt alan adlarını
    yok sayar) - örn. "yatirimci.sirket.com.tr" ve "www.sirket.com.tr" için
    ikisi de "sirket.com.tr" döner, ama "farkli-sirket.com" için farklı olur.
    """
    host = netloc.split(":")[0].lower()
    parts = host.split(".")
    if len(parts) <= 2:
        return host
    if parts[-1] == "tr" and parts[-2] in _TR_SECOND_LEVEL_TLDS:
        return ".".join(parts[-3:])
    return ".".join(parts[-2:])


def _hub_link_score(href: str, anchor_text: str) -> int:
    haystack = _fold_ascii(href) + " " + _fold_ascii(anchor_text)
    return sum(kw in haystack for kw in _HUB_LINK_KEYWORDS)


def _fetch_html(url: str, session: requests.Session, timeout: float) -> str | None:
    try:
        time.sleep(REQUEST_DELAY_SECONDS)
        resp = session.get(url, timeout=timeout, headers=BROWSER_HEADERS)
        resp.raise_for_status()
    except requests.RequestException as exc:
        logger.debug("Sayfa alınamadı (%s): %s", url, exc)
        return None
    return resp.text


def _extract_pdf_links(html: str, base_url: str, same_domain: str) -> set[str]:
    """Sayfadaki .pdf bağlantılarını toplar - yalnızca AYNI (kayıtlı) alan
    adındakiler. Kısıtlama olmasa footer'daki bir ortaklık/haber linkinden
    tamamen alakasız bir üçüncü taraf PDF'i (örn. bir STK'nın raporu) yanlışlıkla
    şirketin raporu sanılabiliyor."""
    links: set[str] = set()
    if BeautifulSoup is not None:
        soup = BeautifulSoup(html, "html.parser")
        hrefs = [a["href"] for a in soup.find_all("a", href=True) if a["href"].lower().split("?")[0].endswith(".pdf")]
    else:  # yedek: basit regex
        hrefs = [m.group(1) for m in _PDF_LINK_RE.finditer(html)]

    for href in hrefs:
        abs_url = urljoin(base_url, href)
        if _registrable_domain(urlparse(abs_url).netloc) == same_domain:
            links.add(abs_url)
    return links


def _extract_hub_links(html: str, base_url: str, same_domain: str) -> list[tuple[int, str]]:
    """Sayfadaki, rapor/yatırımcı ilişkileri ile alakalı görünen (aynı kayıtlı
    alan adındaki - alt alan adları dahil) bağlantıları puanlarıyla döner."""
    if BeautifulSoup is None:
        return []
    hubs: list[tuple[int, str]] = []
    soup = BeautifulSoup(html, "html.parser")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if href.lower().split("?")[0].endswith(".pdf"):
            continue
        abs_url = urljoin(base_url, href)
        if _registrable_domain(urlparse(abs_url).netloc) != same_domain:
            continue
        score = _hub_link_score(href, a.get_text() or "")
        if score > 0:
            hubs.append((score, abs_url))
    return hubs


def discover_pdf_links_from_ir_page(
    ir_url: str,
    session: requests.Session,
    timeout: float = config.HTTP_TIMEOUT_SECONDS,
    max_pages: int = config.IR_CRAWL_MAX_PAGES,
) -> list[str]:
    """IR sitesinden başlayıp rapor/yatırımcı ilişkileri ile alakalı görünen
    bağlantıları izleyerek (en fazla `max_pages` sayfa) .pdf bağlantıları toplar.

    Sadece verilen sayfayı taramak yetersiz kalıyordu: gerçek raporlar hemen
    hemen hiçbir zaman ana sayfada doğrudan linklenmiyor, birkaç tık derinde
    ("Yatırımcı İlişkileri" > "Sürdürülebilirlik" > "Raporlar" gibi) oluyor.
    Bu yüzden en-iyi-önce (best-first) küçük bir tarayıcı kullanılır: her
    sayfada bulunan .pdf bağlantıları biriktirilir, ayrıca rapor/yatırımcı
    ilişkileri anahtar kelimeleriyle eşleşen (aynı alan adındaki) bağlantılar
    puanlarına göre kuyruğa eklenip önce en yüksek puanlılar ziyaret edilir.
    Ağ kapalıysa/sayfa hiç açılmazsa [] döner, akışı durdurmaz.
    """
    base_domain = _registrable_domain(urlparse(ir_url).netloc)
    visited: set[str] = set()
    to_visit: list[tuple[int, str]] = [(1_000, ir_url)]  # başlangıç sayfası her zaman ilk ziyaret edilir
    pdf_links: set[str] = set()

    while to_visit and len(visited) < max_pages:
        to_visit.sort(key=lambda t: t[0], reverse=True)
        _, url = to_visit.pop(0)
        if url in visited:
            continue
        visited.add(url)

        html = _fetch_html(url, session, timeout)
        if html is None:
            continue

        pdf_links.update(_extract_pdf_links(html, url, base_domain))

        if len(visited) < max_pages:
            queued = {u for _, u in to_visit}
            for score, hub_url in _extract_hub_links(html, url, base_domain):
                if hub_url not in visited and hub_url not in queued:
                    to_visit.append((score, hub_url))
                    queued.add(hub_url)

    return sorted(pdf_links)


# Rapor/sürdürülebilirlik ile hiçbir ilgisi olmayan ama IR sayfalarında sıkça
# aynı klasörde/yakında duran, isimlerinde "faaliyet"/"sürdürülebilirlik" gibi
# kelimeler geçmese de yıl/tür skorlamasını yanıltabilen belge türleri: genel
# risk bildirim formları (bankalarda zorunlu, "...faaliyetleri..." içerir),
# politika/sertifika belgeleri (rapor değil, TEK SAYFALIK taahhüt/duyuru).
# Bunlar AKBNK/YKBNK'nin "Faaliyet Raporu" yerine risk bildirim formu, BALSU'nun
# "Sürdürülebilirlik Raporu" yerine politika/ISO sertifikası indirmesine yol
# açmıştı - skorlamaya girmeden tamamen elenirler.
_NEGATIVE_LINK_KEYWORDS = (
    "risk-bildirim", "risk bildirim", "riskbildirim",
    "politika", "sertifika", "sertifikasyon",
    "beyan", "kvkk", "taahhutname", "aydinlatma-metni", "aydinlatma metni",
)


def pick_best_pdf_link(links: list[str], yil: int, rapor_turu: str) -> str | None:
    """Aday PDF bağlantıları arasından yıl + rapor türü ipuçlarına en uygun olanı seçer."""
    yil_str = str(yil)
    # NOT: "TSRS Uyumlu" ve "Gönüllü" aynı belge adayı havuzunu paylaşır - IR
    # taraması ikisini URL/dosya adından güvenilir biçimde ayıramaz (bkz.
    # build_ir_page_refs), bu yüzden ikisi de aynı sürdürülebilirlik anahtar
    # kelimeleriyle aranır; TSRS/Gönüllü ayrımı indirme SONRASI içerik
    # kontrolüyle netleşir (bkz. run()'daki _promote_if_tsrs_content).
    type_keywords = (
        ["faaliyet", "annual"]
        if rapor_turu == config.REPORT_TYPE_FAALIYET
        else ["surdurulebilirlik", "sustainability", "entegre", "integrated"]
    )

    def normalize(s: str) -> str:
        return s.lower().translate(str.maketrans("ışğüöç", "isguoc"))

    scored = []
    for link in links:
        path = urlparse(link).path
        if any(kw in _fold_ascii(path) for kw in _NEGATIVE_LINK_KEYWORDS):
            continue
        name = normalize(path)
        score = (yil_str in name) + sum(kw in name for kw in type_keywords)
        if score > 0:
            scored.append((score, link))

    if not scored:
        return None
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[0][1]


def build_ir_page_refs(
    companies: list[Company],
    ir_cache: dict[str, InvestorRelationsInfo],
    existing_keys: set[tuple[str, int, str]],
    session: requests.Session,
) -> list[ReportRef]:
    """KGK envanterinde eksik kalan şirket-yıl-rapor türü kombinasyonları için IR sayfasını dener.

    Şirket başına IR sitesi TEK SEFER taranır (6 yıl×tür kombinasyonu için
    ayrı ayrı değil) - hem gereksiz tekrar isteği önler hem de o bütçeyle
    `discover_pdf_links_from_ir_page`'in çok sayfalı taramasını (bkz.
    IR_CRAWL_MAX_PAGES) karşılamaya yeter.
    """
    refs: list[ReportRef] = []
    for company in companies:
        # "TSRS Uyumlu" ve "Gönüllü" kategorileri aynı belgeyi (şirketin TEK
        # sürdürülebilirlik raporu) temsil eder - IR taraması bu ikisini ayrı
        # ayrı ARAMAZ, tek bir "Gönüllü" adayı arar (indirme sonrası içerik
        # TSRS'e atıf yapıyorsa run() bunu TSRS Uyumlu'ya yükseltir). Bu yüzden
        # bir yıl için TSRS kategorisi zaten çözülmüşse (KGK ya da önceki bir
        # başarılı indirme ile) o yıl için Gönüllü de aranmaz.
        needed: list[tuple[int, str]] = []
        for yil in config.SCAN_YEARS:
            sustainability_resolved = (
                (company.kod, yil, config.REPORT_TYPE_TSRS) in existing_keys
                or (company.kod, yil, config.REPORT_TYPE_GONULLU) in existing_keys
            )
            if not sustainability_resolved:
                needed.append((yil, config.REPORT_TYPE_GONULLU))
            if (company.kod, yil, config.REPORT_TYPE_FAALIYET) not in existing_keys:
                needed.append((yil, config.REPORT_TYPE_FAALIYET))
        if not needed:
            continue

        ir_info = ir_cache.get(company.kod)
        ir_url = ir_info.yatirimci_iliskileri_url or ir_info.resmi_web_sitesi if ir_info else None

        links = discover_pdf_links_from_ir_page(ir_url, session) if ir_url else []
        # Aynı PDF, aynı şirket için birden fazla yıl/tür kombinasyonuna
        # ATANMAZ: bir belge aynı anda hem 2023 hem 2024 hem de hem Faaliyet
        # hem Sürdürülebilirlik raporu olamaz. IR sayfasında tek bir PDF
        # bulunduğunda (örn. küçük bir holding grubunun ortak sitesi) eskiden
        # bu tek link 6 kombinasyona da (hatta grup şirketlerinin hepsine)
        # kopyalanıyordu - bkz. PAHOL/PASEU/PSGYO'nun hepsinin aynı Pasifik
        # Holding ara dönem faaliyet raporunu "indirmiş" görünmesi hatası.
        used_links: set[str] = set()
        for yil, rapor_turu in needed:
            available = [link for link in links if link not in used_links]
            best = pick_best_pdf_link(available, yil, rapor_turu) if available else None
            if best:
                used_links.add(best)
                refs.append(
                    ReportRef(
                        kod=company.kod,
                        yil=yil,
                        rapor_turu=rapor_turu,
                        kaynak_url=best,
                        kaynak_turu="Şirket IR Sitesi",
                        durum=ReportStatus.INDIRILECEK,
                    )
                )
            else:
                refs.append(
                    ReportRef(kod=company.kod, yil=yil, rapor_turu=rapor_turu, durum=ReportStatus.BULUNMADI)
                )
    return refs


# --------------------------------------------------------------------------
# İndirme (idempotent)
# --------------------------------------------------------------------------

def _slug(text: str) -> str:
    return re.sub(r"[^a-zA-Z0-9]+", "_", text).strip("_")


def local_path_for(kod: str, yil: int, rapor_turu: str) -> Path:
    return config.DOWNLOADS_DIR / kod / str(yil) / f"{_slug(rapor_turu)}.pdf"


def _is_valid_pdf(path: Path) -> bool:
    if not path.exists() or path.stat().st_size < config.HTTP_MIN_PDF_BYTES:
        return False
    with path.open("rb") as f:
        return f.read(5) == b"%PDF-"


# TSRS uygunluk kontrolü: rapor gerçekten "Türkiye Sürdürülebilirlik Raporlama
# Standartları"na uygun mu, yoksa şirket sitesinden bulunan alakasız bir PDF mi
# (örn. bir politika belgesi, sertifika) - akademik makalenin kapsamı için bu
# ayrım kritik, özellikle KGK dışı (IR sitesi) kaynaklı indirmelerde.
_TSRS_MARKERS_FOLDED = ("tsrs", "turkiye surdurulebilirlik raporlama standartlari")


def check_tsrs_compliance(path: Path, max_pages: int = 3) -> bool | None:
    """PDF'in ilk `max_pages` sayfasında "TSRS" ya da açık adı geçiyor mu kontrol eder.

    PyMuPDF kurulu değilse ya da PDF açılamazsa None (bilinmiyor) döner -
    "Hayır" ile karıştırılmamalı, bu yüzden Excel'de ayrı gösterilir.
    """
    if fitz is None:
        return None
    try:
        with fitz.open(path) as doc:
            text = "".join(page.get_text() for page in list(doc)[:max_pages])
    except Exception as exc:  # noqa: BLE001 - bozuk/şifreli PDF akışı durdurmasın
        logger.debug("TSRS kontrolü için PDF açılamadı (%s): %s", path, exc)
        return None
    folded = _fold_ascii(text)
    return any(marker in folded for marker in _TSRS_MARKERS_FOLDED)


# Genel yasal risk bildirim/politika/aydınlatma belgeleri: bunlar rapor DEĞİL,
# ama IR sayfalarında yıllık raporlarla aynı klasörde/yakında durabiliyor ve
# dosya adları yıl/tür skorlamasını (pick_best_pdf_link) yanıltabiliyor - bkz.
# AKBNK/YKBNK'nin "Faaliyet Raporu" yerine genel risk bildirim formunu
# "indirmiş" görünmesi hatası. URL bazlı eleme (_NEGATIVE_LINK_KEYWORDS) her
# zaman yeterli değil (dosya adı temiz olabilir) - bu yüzden İÇERİK bazlı ikinci
# bir savunma katmanı: gerçek faaliyet/sürdürülebilirlik raporları neredeyse
# hiçbir zaman bu kadar KISA olmaz (genelde onlarca-yüzlerce sayfa), bu tür
# genel bildirim belgeleri ise tipik olarak birkaç sayfalık tek konulu
# metinlerdir. Bu yüzden hem "kısa belge" hem "bu ifadelerden biri geçiyor"
# koşulu birlikte aranır - aksi halde gerçek bir raporun "politika" kelimesini
# bir bölüm başlığında geçirmesi (çok sayfalı raporlarda sık) yanlışlıkla
# reddedilir.
_GENERIC_DISCLOSURE_KEYWORDS_FOLDED = (
    # "risk bildirim" (gövde/stem) hem "Risk Bildirimi" hem "Risk Bildirim
    # Formu" gibi varyasyonları yakalar - kullanıcı talebindeki tam ifade
    # ("risk bildirimi") yalnızca ilkini yakalardı.
    "risk bildirim", "politika", "aydinlatma metni", "genel bilgilendirme",
)
_GENERIC_DISCLOSURE_MAX_PAGES = 12


def check_generic_disclosure(path: Path, max_pages: int = 3) -> bool | None:
    """PDF gerçek bir yıllık rapor değil de genel bir risk bildirimi/politika/
    aydınlatma metni gibi mi görünüyor kontrol eder (bkz. yukarıdaki not).

    True dönerse çağıran taraf indirmeyi REDDETMELİDİR. PyMuPDF kurulu değilse
    ya da PDF açılamazsa None (bilinmiyor, reddetme) döner.
    """
    if fitz is None:
        return None
    try:
        with fitz.open(path) as doc:
            if doc.page_count > _GENERIC_DISCLOSURE_MAX_PAGES:
                return False
            text = "".join(page.get_text() for page in list(doc)[:max_pages])
    except Exception as exc:  # noqa: BLE001 - bozuk/şifreli PDF akışı durdurmasın
        logger.debug("Genel bildirim kontrolü için PDF açılamadı (%s): %s", path, exc)
        return None
    folded = _fold_ascii(text)
    return any(kw in folded for kw in _GENERIC_DISCLOSURE_KEYWORDS_FOLDED)


def _build_sibling_words(
    companies_by_kod: dict[str, Company], ir_cache: dict[str, InvestorRelationsInfo]
) -> dict[str, set[str]]:
    """Aynı (kayıtlı) alan adını paylaşan kardeş şirketlerin ünvanlarında
    ORTAK geçen kelimeleri, şirket kodu bazında döner.

    Bir holding grubunun tüm üyeleri genelde ortak yatırımcı ilişkileri
    sitesini paylaşır (örn. pasifik.com -> PAHOL/PASEU/PSGYO) ve grup markası
    ünvanlarının hepsinde geçer (örn. "PASİFİK"). Bu kelime örneklemin
    tamamında (_build_generic_unvan_words eşiği: >=5 şirket) genel sayılacak
    kadar sık geçmeyebilir - ama TAM DA bu grup içinde ayırt edici değildir,
    bu yüzden check_company_match'te ayrıca (grup-yerel) genel kelime olarak
    ele alınmalı; aksi halde grup markası kelimesi eşleşmesi, kardeş
    şirketin raporunun yanlışlıkla doğrulanmasına yol açar (bkz. PASEU'nun
    PAHOL'ün raporunu "kendi raporuymuş gibi" geçirmesi).
    """
    domain_to_kods: dict[str, list[str]] = {}
    for kod, info in ir_cache.items():
        url = info.yatirimci_iliskileri_url or info.resmi_web_sitesi
        if not url:
            continue
        domain = _registrable_domain(urlparse(url).netloc)
        domain_to_kods.setdefault(domain, []).append(kod)

    sibling_words_by_kod: dict[str, set[str]] = {}
    for kods in domain_to_kods.values():
        if len(kods) < 2:
            continue
        word_counts: dict[str, int] = {}
        for kod in kods:
            company = companies_by_kod.get(kod)
            if company is None:
                continue
            for w in _unvan_words(company.unvan):
                word_counts[w] = word_counts.get(w, 0) + 1
        shared = {w for w, n in word_counts.items() if n >= 2}
        if not shared:
            continue
        for kod in kods:
            sibling_words_by_kod[kod] = shared
    return sibling_words_by_kod


def check_company_match(path: Path, company: Company, exclude_words: set[str], max_pages: int = 1) -> bool | None:
    """IR sitesinden indirilen PDF gerçekten BU şirkete mi ait, kontrol eder.

    IR sayfası taraması (discover_pdf_links_from_ir_page) sadece aynı alan
    adına bakar - ama bir holding grubunun ortak yatırımcı ilişkileri sitesi
    (örn. pasifik.com), grup içindeki KARDEŞ şirketler için de aynı tek PDF'i
    döndürebilir (bkz. PAHOL/PASEU/PSGYO: hepsi Pasifik Holding'in kendi
    faaliyet raporunu "indirmiş" görünüyordu). KGK kaynaklı indirmeler zaten
    load_tsrs_envanteri_refs'te başlık bazlı bu kontrolden geçiyor; burada
    aynı mantık (_unvan_words/_has_shared_word), PDF'in gerçek metnine karşı
    uygulanır. `exclude_words` hem örneklem-geneli genel kelimeleri hem de
    (varsa) aynı IR sitesini paylaşan kardeş şirketlerin ortak marka
    kelimesini içermeli (bkz. _build_sibling_words) - aksi halde grup markası
    eşleşmesi tek başına yeterli sayılıp kardeş şirketin raporu yanlışlıkla
    doğrulanır. Şirketin ünvanı bunlar çıkarıldıktan sonra hiç ayırt edici
    kelime bırakmıyorsa karar verilemez (None) - yanlışlıkla reddetmemek için.

    max_pages varsayılan olarak SADECE kapak sayfasına (1) bakar - bir holding
    raporunun gövdesinde genelde tüm grup şirketleri (bağlı ortaklıklar
    listesi/organizasyon şeması) adı geçtiğinden, birkaç sayfa taransa kardeş
    şirketin adı da metinde "bulunur" ve kontrol yanlışlıkla doğrular; kapak
    sayfası ise raporun KİME ait olduğunu tekil ve güvenilir biçimde belirtir.
    """
    company_words = _unvan_words(company.unvan) - exclude_words
    if not company_words:
        return None
    if fitz is None:
        return None
    try:
        with fitz.open(path) as doc:
            text = "".join(page.get_text() for page in list(doc)[:max_pages])
    except Exception as exc:  # noqa: BLE001 - bozuk/şifreli PDF akışı durdurmasın
        logger.debug("Şirket eşleşme kontrolü için PDF açılamadı (%s): %s", path, exc)
        return None
    doc_words = _title_words(text)
    return _has_shared_word(company_words, doc_words)


def download_report(
    ref: ReportRef,
    session: requests.Session,
    company: Company | None = None,
    exclude_words: set[str] | None = None,
) -> ReportRef:
    """Tek bir raporu indirir. İdempotent: geçerli bir dosya zaten varsa atlar.

    KGK kaynaklı istekler için daha az sabırlı bir politika uygulanır: bu
    ortamdan kgk.gov.tr'ye giden isteklerin neredeyse tamamı 500 ile
    sonuçlanıyor - 3 deneme × üstel geri çekilme (~15sn/rapor) sadece zaman
    kaybettiriyor. KGK_MAX_RETRIES/KGK_TIMEOUT_SECONDS (bkz. config.py) daha
    düşük tutularak hem gerçekten çalıştığı durumlar (geçici WAF engeli)
    yakalanır hem de toplam süre önemli ölçüde kısalır.

    `company`/`exclude_words` verildiğinde ve kaynak "Şirket IR Sitesi" ise,
    indirilen (ya da diskte zaten var olan) dosyanın gerçekten bu şirkete ait
    olup olmadığı da doğrulanır (bkz. check_company_match) - IR sayfası
    taraması bir holding grubunun ortak sitesinden kardeş şirkete ait yanlış
    bir raporu bulmuşsa, dosya silinip "Hata" olarak işaretlenir; sessizce
    "İndirildi" görünmesi engellenir. Kaynaktan BAĞIMSIZ olarak, indirilen
    (ya da diskte zaten var olan) her PDF ayrıca gerçek bir yıllık rapor mu
    yoksa genel bir risk bildirimi/politika/aydınlatma metni mi diye de
    kontrol edilir (bkz. check_generic_disclosure) - bkz. AKBNK/YKBNK'nin
    "Faaliyet Raporu" yerine risk bildirim formu indirmiş görünmesi hatası.
    """
    dest = local_path_for(ref.kod, ref.yil, ref.rapor_turu)

    def _reject_company_mismatch() -> None:
        dest.unlink(missing_ok=True)
        ref.yerel_dosya = None
        ref.durum = ReportStatus.HATA
        ref.hata_mesaji = (
            "İndirilen belge şirket ünvanıyla eşleşmiyor "
            "(muhtemel çapraz şirket/holding karışıklığı)"
        )

    def _reject_generic_disclosure() -> None:
        dest.unlink(missing_ok=True)
        ref.yerel_dosya = None
        ref.durum = ReportStatus.HATA
        ref.hata_mesaji = (
            "İndirilen belge gerçek bir yıllık rapor değil, genel bir risk "
            "bildirimi/politika/aydınlatma metni gibi görünüyor (içerik "
            "doğrulaması reddetti)"
        )

    def _company_mismatch() -> bool:
        if company is None or exclude_words is None or ref.kaynak_turu != "Şirket IR Sitesi":
            return False
        return check_company_match(dest, company, exclude_words) is False

    def _generic_disclosure() -> bool:
        return check_generic_disclosure(dest) is True

    if _is_valid_pdf(dest):
        if _company_mismatch():
            _reject_company_mismatch()
            return ref
        if _generic_disclosure():
            _reject_generic_disclosure()
            return ref
        ref.yerel_dosya = str(dest)
        ref.durum = ReportStatus.INDIRILDI
        ref.tsrs_uyumlu = check_tsrs_compliance(dest)
        return ref

    if not ref.kaynak_url:
        ref.durum = ReportStatus.BULUNMADI
        return ref

    is_kgk = ref.kaynak_turu == "KGK"
    max_retries = config.KGK_MAX_RETRIES if is_kgk else config.HTTP_MAX_RETRIES
    timeout = config.KGK_TIMEOUT_SECONDS if is_kgk else config.HTTP_TIMEOUT_SECONDS

    dest.parent.mkdir(parents=True, exist_ok=True)
    last_exc: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            time.sleep(REQUEST_DELAY_SECONDS)
            resp = session.get(
                ref.kaynak_url,
                timeout=timeout,
                headers=BROWSER_HEADERS,
                stream=True,
            )
            resp.raise_for_status()
            tmp_dest = dest.with_suffix(".part")
            with tmp_dest.open("wb") as f:
                for chunk in resp.iter_content(chunk_size=65536):
                    f.write(chunk)
            tmp_dest.replace(dest)

            if not _is_valid_pdf(dest):
                raise ValueError("İndirilen dosya geçerli bir PDF değil (magic bytes/boyut uyuşmuyor)")

            if _company_mismatch():
                _reject_company_mismatch()
                return ref
            if _generic_disclosure():
                _reject_generic_disclosure()
                return ref

            ref.yerel_dosya = str(dest)
            ref.durum = ReportStatus.INDIRILDI
            ref.tsrs_uyumlu = check_tsrs_compliance(dest)
            return ref
        except (requests.RequestException, ValueError, OSError) as exc:
            last_exc = exc
            logger.warning(
                "İndirme denemesi %d/%d başarısız (%s, %s, %s): %s",
                attempt, max_retries, ref.kod, ref.yil, ref.rapor_turu, exc,
            )
            if attempt < max_retries:
                time.sleep(config.HTTP_RETRY_BACKOFF_SECONDS * attempt)

    ref.durum = ReportStatus.HATA
    ref.hata_mesaji = str(last_exc)
    return ref


_LEGACY_RAPOR_TURU_MAP = {
    # Eski (2'li) REPORT_TYPES şemasından kalma etiket -> yeni (3'lü) karşılığı.
    # "Faaliyet Raporu" değişmedi, eşleme gerekmez.
    "Sürdürülebilirlik Raporu": config.REPORT_TYPE_GONULLU,
}


def _migrate_legacy_ref(ref: ReportRef) -> ReportRef:
    """Eski şemadan kalma bir kaydı yeni 3'lü REPORT_TYPES şemasına taşır.

    Taşınmazsa bu kayıt YENİ etiketlerden (config.REPORT_TYPES) hiçbiriyle
    eşleşmeyen bir .key altında sonsuza dek "yetim" kalır: excel_writer.py
    yalnızca 3 yeni etiketi sorguladığı için böyle bir kayıt ASLA Excel'de
    görünmez (sessiz veri kaybı), ayrıca _promote_if_tsrs_content de eski
    etiketi tanımadığı için TSRS'e hiç yükseltemez. Yerel dosya (varsa) YENİ
    kategorinin dosya yoluna taşınır ki idempotent kısayol onu bulabilsin.
    """
    new_type = _LEGACY_RAPOR_TURU_MAP.get(ref.rapor_turu)
    if new_type is None:
        return ref
    migrated = replace(ref, rapor_turu=new_type)
    if migrated.yerel_dosya:
        old_path = Path(migrated.yerel_dosya)
        if old_path.exists():
            new_path = local_path_for(migrated.kod, migrated.yil, migrated.rapor_turu)
            new_path.parent.mkdir(parents=True, exist_ok=True)
            old_path.replace(new_path)
            migrated.yerel_dosya = str(new_path)
    return migrated


def _promote_if_tsrs_content(ref: ReportRef) -> ReportRef:
    """"Gönüllü" olarak indirilen bir raporu, içeriği TSRS'e atıf yapıyorsa
    "TSRS Uyumlu"ya yükseltir (bkz. config.py'deki 3'lü REPORT_TYPES notu:
    "... VEYA ilk 3 sayfada TSRS standardına atıf yapan").

    rapor_turu değişince ReportRef.key de değişir - yerel dosya, YENİ
    kategorinin dosya yoluna TAŞINIR ki bir sonraki çalıştırmada idempotent
    kısayol onu doğru yerde bulsun (aksi halde her çalıştırmada gereksiz
    yeniden indirme tetiklenir). Çağıran taraf (run()) refs_by_key'deki eski
    anahtarı silip yeni anahtarla değiştirmekten sorumludur.
    """
    if ref.durum != ReportStatus.INDIRILDI or ref.rapor_turu != config.REPORT_TYPE_GONULLU or not ref.tsrs_uyumlu:
        return ref
    old_path = Path(ref.yerel_dosya) if ref.yerel_dosya else None
    ref.rapor_turu = config.REPORT_TYPE_TSRS
    if old_path and old_path.exists():
        new_path = local_path_for(ref.kod, ref.yil, ref.rapor_turu)
        new_path.parent.mkdir(parents=True, exist_ok=True)
        old_path.replace(new_path)
        ref.yerel_dosya = str(new_path)
    return ref


def _enforce_unique_urls(refs_by_key: dict[tuple, ReportRef], companies_by_kod: dict[str, Company]) -> int:
    """Aynı kaynak_url'nin birden fazla BIST koduna atanmasını KESİN olarak
    engeller (bkz. PAHOL/PASEU/PSGYO'nun hepsinin aynı Pasifik Holding PDF'ini
    "indirmiş" görünmesi hatası).

    check_company_match sezgiseldir (ünvan kelimesi eşleşmesine dayanır) ve
    örneklemin bazı köşe durumlarında (örn. ünvanın ayırt edici hiçbir
    kelimesi kalmaması) None (bilinmiyor) dönebilir - bu fonksiyon buna
    GÜVENMEZ, tamamen ayrı ve kesin bir son-kontrol katmanıdır: aynı URL
    birden fazla FARKLI şirket koduna "İndirildi" olarak atanmışsa, içerik
    eşleşmesi doğrulanabilen TEK bir kod (varsa) kazanır, diğerlerinin
    yerel dosyaları SİLİNİR ve "Rapor Yok"a düşürülür. Hiçbiri için eşleşme
    doğrulanamıyorsa (ör. fitz yok) TÜMÜ "Rapor Yok"a düşürülür - belirsiz bir
    raporu rastgele bir şirkete atfetmektense hiç atfetmemek tercih edilir.
    """
    url_to_keys: dict[str, list[tuple]] = {}
    for key, ref in refs_by_key.items():
        if ref.kaynak_url and ref.durum == ReportStatus.INDIRILDI:
            url_to_keys.setdefault(ref.kaynak_url, []).append(key)

    downgraded = 0
    for url, keys in url_to_keys.items():
        kods = {k[0] for k in keys}
        if len(kods) <= 1:
            continue  # aynı şirketin farklı yıl/tür kombinasyonları için sorun değil

        best_key = None
        for key in keys:
            ref = refs_by_key[key]
            company = companies_by_kod.get(ref.kod)
            if not company or not ref.yerel_dosya:
                continue
            if check_company_match(Path(ref.yerel_dosya), company, set()) is True:
                best_key = key
                break

        for key in keys:
            if key == best_key:
                continue
            ref = refs_by_key[key]
            logger.warning(
                "Çapraz şirket çakışması: kaynak_url (%s) %s içinde birden "
                "fazla BIST koduna atanmıştı - %s 'Rapor Yok'a düşürüldü.",
                url, sorted(kods), key,
            )
            if ref.yerel_dosya:
                Path(ref.yerel_dosya).unlink(missing_ok=True)
            ref.yerel_dosya = None
            ref.durum = ReportStatus.BULUNMADI
            ref.hata_mesaji = None
            ref.kaynak_url = None
            ref.tsrs_uyumlu = None
            downgraded += 1
    return downgraded


# --------------------------------------------------------------------------
# Manifest (text_miner.py / excel_writer.py bunu okur)
# --------------------------------------------------------------------------

def load_manifest(path=config.REPORT_MANIFEST_FILE) -> list[ReportRef]:
    raw = load_json(path, [])
    refs = []
    for item in raw:
        item = dict(item)
        item["durum"] = ReportStatus(item["durum"])
        refs.append(ReportRef(**item))
    return refs


def save_manifest(refs: list[ReportRef], path=config.REPORT_MANIFEST_FILE) -> None:
    save_json(path, [{**asdict(r), "durum": r.durum.value} for r in refs])


# --------------------------------------------------------------------------
# Parça bazlı çalıştırma
# --------------------------------------------------------------------------

def run(
    companies: list[Company],
    progress_store: ProgressStore,
    chunk_size: int = config.CHUNK_SIZE,
) -> list[ReportRef]:
    companies_by_kod = {c.kod: c for c in companies}
    generic_unvan_words = _build_generic_unvan_words(companies_by_kod)
    ir_cache = load_ir_cache()
    sibling_words_by_kod = _build_sibling_words(companies_by_kod, ir_cache)
    session = requests.Session()

    # Önceki başarılı indirmeleri kaybetmemek için: KGK gibi "tercih edilen"
    # bir kaynak bu çalıştırmada mevcut olduğunda IR taraması o kombinasyon
    # için hiç denenmez (existing_keys) - KGK indirmesi başarısız olursa
    # (kgk.gov.tr çoğu istekte 500 döndürüyor) kombinasyon "Hata"ya düşüp
    # önceki bir çalıştırmada IR sitesinden başarıyla indirilmiş geçerli bir
    # PDF varken kaybedilmemeli. Aşağıda tüm indirme denemeleri bittikten
    # sonra bu türden regresyonlar önceki başarılı sonuçla geri doldurulur.
    previous_manifest_by_key: dict[tuple, ReportRef] = {}
    for _r in load_manifest():
        _r = _migrate_legacy_ref(_r)
        previous_manifest_by_key[_r.key] = _r
    previous_success_by_key = {
        key: r for key, r in previous_manifest_by_key.items()
        if r.durum == ReportStatus.INDIRILDI and r.yerel_dosya
    }

    if config.ENABLE_KGK_SOURCE:
        tsrs_refs = load_tsrs_envanteri_refs(companies_by_kod)
    else:
        logger.info("KGK kaynağı devre dışı (config.ENABLE_KGK_SOURCE=False) - hiç denenmiyor.")
        tsrs_refs = []
    # IR taraması yalnızca KGK'nın kapsadığı kombinasyonları değil, önceki bir
    # çalıştırmada ZATEN başarıyla indirilmiş kombinasyonları da atlamalı -
    # aksi halde her yeniden çalıştırma, dosyası zaten diskte olan onlarca
    # şirket için (gereksiz yere) IR sitesini baştan tarar (bkz.
    # discover_pdf_links_from_ir_page - şirket başına birkaç sayfalık ağ
    # isteği demek).
    existing_keys = {r.key for r in tsrs_refs} | set(previous_success_by_key)
    ir_refs = build_ir_page_refs(companies, ir_cache, existing_keys, session)

    all_refs = tsrs_refs + ir_refs
    refs_by_key = {r.key: r for r in all_refs}
    for key, prev in previous_success_by_key.items():
        # KGK bu anahtarı da kapsıyorsa onun denemesi önceliklidir (başarısız
        # olursa regresyon-önleme adımı zaten prev'i geri yükleyecek);
        # kapsamıyorsa (IR taraması da atladığı için) prev burada eklenmezse
        # manifestten tamamen kaybolur. replace() ile KOPYALANIR - prev
        # nesnesinin kendisi eklenirse, process() onu daha sonra reddedip
        # (örn. yerel_dosya=None) mutasyona uğratabilir; bu da previous_
        # success_by_key/previous_manifest_by_key İÇİNDEKİ AYNI nesneyi de
        # sessizce bozar (bkz. aşağıdaki regressed döngüsünün Path(None) ile
        # çökmesi hatası - iki sözlük aynı nesneyi paylaştığı için oluştu).
        refs_by_key.setdefault(key, replace(prev))

    def _attempt(ref: ReportRef) -> ReportRef:
        # check_company_match yalnızca AYNI kayıtlı alan adını paylaşan
        # kardeş şirketler varsa (sibling_words_by_kod'da girişi varsa)
        # çalıştırılır - aksi halde None geçilir ve _company_mismatch()
        # devre dışı kalır (bkz. download_report). Sebebi: kontrol sadece
        # KAPAK SAYFASINA bakıyor (bkz. check_company_match docstring'i) ve
        # birçok gerçek rapor kapağı tasarım gereği şirket adını metin
        # olarak İÇERMEZ (logo/görsel) - örneklemde bunu ölçtüğümüzde
        # (AEFES/ARCLK/BIMAS/PETKM/SAHOL/TSKB/TTKOM gibi tamamen kendi
        # alan adından indirilen, çapraz-şirket riski SIFIR olan raporlarda)
        # bu yanlışlıkla "eşleşmiyor" diye reddediyordu. Çapraz-şirket riski
        # yalnızca aynı alan adını paylaşan gruplarda (örn. PAHOL/PASEU/
        # PSGYO) var - başka hiçbir şirket için bu URL zaten ortaya
        # çıkamaz, bu yüzden kontrol de yalnızca oralarda gerekli. Kesin
        # çapraz-atama garantisi zaten _enforce_unique_urls'ten gelir.
        sibling_words = sibling_words_by_kod.get(ref.kod)
        exclude_words = (generic_unvan_words | sibling_words) if sibling_words else None
        updated = download_report(ref, session, companies_by_kod.get(ref.kod), exclude_words)
        return _promote_if_tsrs_content(updated)

    def process(ref: ReportRef) -> None:
        old_key = ref.key
        updated = _attempt(ref)
        new_key = updated.key
        if new_key != old_key:
            refs_by_key.pop(old_key, None)
        refs_by_key[new_key] = updated

    def on_chunk_done(chunk_idx: int, total_chunks: int, progress) -> None:
        save_manifest(list(refs_by_key.values()))
        indirildi = sum(1 for r in refs_by_key.values() if r.durum == ReportStatus.INDIRILDI)
        rapor_yok = sum(1 for r in refs_by_key.values() if r.durum == ReportStatus.BULUNMADI)
        hata = sum(1 for r in refs_by_key.values() if r.durum == ReportStatus.HATA)
        print(
            f"[downloader] Parça {chunk_idx}/{total_chunks} — "
            f"indirildi: {indirildi}, rapor yok: {rapor_yok}, hata: {hata}"
        )

    # Sadece indirilecek olanları (URL'si olanları) chunk'lı akışa sokuyoruz;
    # "Rapor Yok" olanlar zaten nihai durumda, boşuna işlenmez. refs_by_key
    # (all_refs değil) kullanılır ki setdefault ile eklenen "önceki başarılı
    # indirme" kayıtları da download_report'un ucuz (ağ gerektirmeyen, sadece
    # yerel dosyayı doğrulayan) idempotent kısayoluna girsin - bu sayede yeni
    # eklenen TSRS uygunluk kontrolü onlar için de geriye dönük hesaplanır.
    to_download = [r for r in refs_by_key.values() if r.kaynak_url]

    run_in_chunks(
        items=to_download,
        key_fn=lambda r: "|".join(map(str, r.key)),
        process_fn=process,
        progress_store=progress_store,
        stage_name="download",
        chunk_size=chunk_size,
        on_chunk_done=on_chunk_done,
    )

    # Bir kombinasyon önceki çalıştırmada zaten denenmişse (progress.json'da
    # "tamamlandı" işaretli), run_in_chunks onu bu turda hiç işlemeden
    # atlar - ama refs_by_key'deki karşılığı yukarıda SIFIRDAN inşa edilen
    # taze bir "İndirilecek" yer tutucusudur (bkz. load_tsrs_envanteri_refs/
    # build_ir_page_refs). Önceki deneme başarılıysa (İndirildi) bu zaten
    # previous_success_by_key/setdefault ile yukarıda taşınmıştı; ama önceki
    # deneme "Hata" ile sonuçlanmışsa hiçbir yere taşınmıyordu ve bu kayıt
    # sessizce "İndirilecek"te asılı kalıp hem özet sayımlarından hem de
    # excel_writer'ın "Eksik Raporlar" sekmesinden düşüyordu. Burada önceki
    # manifestteki GERÇEK son durumu (Hata dahil) geri yüklüyoruz.
    stuck_restored = 0
    for key, ref in refs_by_key.items():
        if ref.durum == ReportStatus.INDIRILECEK:
            prev = previous_manifest_by_key.get(key)
            if prev is not None:
                refs_by_key[key] = replace(prev)  # kopya - bkz. yukarıdaki setdefault notu
                stuck_restored += 1
    if stuck_restored:
        logger.info(
            "%d kombinasyon progress.json'da 'tamamlandı' işaretliydi ama bu "
            "çalıştırmada atlandığı için taze bir 'İndirilecek' yer "
            "tutucusunda kalmıştı; önceki çalıştırmanın gerçek sonucu "
            "(İndirildi/Hata) geri yüklendi.",
            stuck_restored,
        )

    regressed = 0
    for key, ref in list(refs_by_key.items()):
        if ref.durum != ReportStatus.INDIRILDI:
            prev = previous_success_by_key.get(key)
            if prev and prev.yerel_dosya and _is_valid_pdf(Path(prev.yerel_dosya)):
                refs_by_key[key] = replace(prev)  # kopya - bkz. yukarıdaki setdefault notu
                regressed += 1
    if regressed:
        logger.info(
            "%d kombinasyon için bu çalıştırmadaki deneme başarısız oldu, "
            "önceki başarılı indirme korundu (regresyon önlendi).",
            regressed,
        )
    if stuck_restored or regressed:
        save_manifest(list(refs_by_key.values()))  # aşağıdaki (uzun sürebilen) ikincil deneme kesilirse bu iş kaybolmasın

    # KGK ikincil deneme: existing_keys yukarıda IR taramasını yalnızca KGK'nın
    # bir ADAY URL ÖNERDİĞİ kombinasyonlar için atlıyordu - "önerdi" GERÇEKTEN
    # indirilebildi anlamına gelmiyor. kgk.gov.tr bu ortamdan gelen isteklerin
    # neredeyse tamamına 500 döndürdüğü için (bkz. modül başı NOT), KGK'nın
    # başlıktan doğru eşleştirdiği ama indiremediği kombinasyonlar (örn.
    # TCELL/THYAO/TUPRS/SISE/ULKER gibi büyük/şeffaf şirketlerin GERÇEK yıllık
    # raporu kendi sitesinde apaçık dururken) hiç denenmeden sürekli "Hata"da
    # kalıyordu. Burada bu şirketler için IR sitesi İKİNCİ bir şans olarak
    # denenir - yalnızca ZATEN başarıyla indirilmiş ya da bu çalıştırmada
    # gerçekten aranıp "Rapor Yok" bulunmuş kombinasyonlar hariç tutulur.
    kgk_failed_companies = sorted({
        ref.kod for ref in refs_by_key.values()
        if ref.durum == ReportStatus.HATA and ref.kaynak_turu == "KGK"
    })
    if kgk_failed_companies:
        logger.info(
            "%d şirket için KGK kaynaklı indirme başarısız oldu; IR sitesi "
            "ikincil deneme olarak taranıyor: %s",
            len(kgk_failed_companies), ", ".join(kgk_failed_companies),
        )
        # Şirket şirket (TEK bir build_ir_page_refs(hepsi) çağrısı yerine) işlenir
        # ve HER şirketten sonra manifest kaydedilir - bu aşama şirket başına
        # birkaç sayfalık ağ taraması içerdiğinden uzun sürebiliyor; bu ortamda
        # bellek baskısı nedeniyle arka plan süreçleri bazen dıştan öldürülüyor
        # (bkz. çalışma günlüğü) - tek seferlik dev bir çağrı olursa süreç
        # ortasında kesilirse TÜM ikincil deneme turu kaybolur. Şirket başına
        # kaydetmek, bir kesilmenin en fazla o an işlenmekte olan TEK şirketin
        # ilerlemesini kaybetmesini sağlar.
        fallback_attempted = 0
        fallback_recovered = 0
        for kod in kgk_failed_companies:
            company = companies_by_kod.get(kod)
            if company is None:
                continue
            fallback_existing_keys = {
                key for key, ref in refs_by_key.items()
                if ref.durum in (ReportStatus.INDIRILDI, ReportStatus.BULUNMADI)
            }
            fallback_refs = build_ir_page_refs([company], ir_cache, fallback_existing_keys, session)
            for fref in fallback_refs:
                if fref.durum != ReportStatus.INDIRILECEK or not fref.kaynak_url:
                    continue  # IR'de de bulunamadı - KGK'nın "Hata"sı zaten yerinde kalsın
                fallback_attempted += 1
                old_key = fref.key
                updated = _attempt(fref)
                if updated.durum == ReportStatus.INDIRILDI:
                    fallback_recovered += 1
                new_key = updated.key
                if new_key != old_key:
                    refs_by_key.pop(old_key, None)
                refs_by_key[new_key] = updated
            save_manifest(list(refs_by_key.values()))
        logger.info(
            "IR ikincil deneme: %d kombinasyon denendi, %d tanesi kurtarıldı.",
            fallback_attempted, fallback_recovered,
        )

    downgraded_shared = _enforce_unique_urls(refs_by_key, companies_by_kod)
    if downgraded_shared:
        logger.info(
            "%d kombinasyon, aynı kaynak_url'yi başka bir BIST koduyla "
            "paylaştığı için 'Rapor Yok'a düşürüldü (bkz. _enforce_unique_urls).",
            downgraded_shared,
        )

    save_manifest(list(refs_by_key.values()))
    return list(refs_by_key.values())


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    from src.inventory import load_bist100

    companies = load_bist100()
    progress_store = ProgressStore(config.PROGRESS_FILE)
    run(companies, progress_store)


if __name__ == "__main__":
    main()

"""Şirket listesini okur; her BIST 100 şirketi için resmi web sitesi ve
yatırımcı ilişkileri (IR) URL'sini bulur/kaydeder (cache'lenir).

Girdi:
  - config.BIST100_LISTE_XLSX   : SABİT 100 şirketlik örneklem (BIST_Kod, Sirket_Unvani, Sehir)
  - config.SIRKETLER_REFERANS_XLSX : BIST'te işlem gören TÜM şirketler (Kod, Ünvan, Şehir,
        Bağımsız Denetim Kuruluşu) — sadece KOD/ünvan eşleştirme REFERANSI, örneklem değil.
  - config.MANUAL_IR_URLS_JSON  : (opsiyonel) araştırmacının elle doldurduğu IR URL'leri.
        Format: {"AKBNK": {"resmi_web_sitesi": "...", "yatirimci_iliskileri_url": "..."}, ...}

Çıktı:
  - config.IR_URL_CACHE_FILE    : kod -> InvestorRelationsInfo cache'i (JSON)

NOT (ağ erişimi): Bu betiğin otomatik-keşif adımı (`probe_url`) canlı bir HTTP
isteği gerektirir. Ajan bazı ortamlarda (ör. bu geliştirme kutusu) dış ağa
kapalı çalıştırılabilir — böyle durumlarda `discover_for_company` tüm adaylar
için başarısız döner ve şirketi "manuel araştırma gerekir" olarak işaretler;
TÜM AKIŞI DURDURMAZ.
"""
from __future__ import annotations

import base64
import json
import logging
import re
import unicodedata
from collections.abc import Callable
from dataclasses import asdict
from urllib.parse import parse_qs, unquote, urlparse

import openpyxl
import requests

import config
from src.models import Company, InvestorRelationsInfo
from src.progress import ProgressStore, run_in_chunks

logger = logging.getLogger(__name__)

try:
    from bs4 import BeautifulSoup
except ImportError:  # pragma: no cover - requirements.txt'de var, savunma amaçlı
    BeautifulSoup = None

SearchFn = Callable[[Company], str | None]  # şirket -> muhtemel resmi site URL'si (varsa)


# --------------------------------------------------------------------------
# Girdi dosyalarını okuma
# --------------------------------------------------------------------------

def load_bist100(path=config.BIST100_LISTE_XLSX) -> list[Company]:
    """SABİT BIST 100 örneklemini okur. Bu liste genişletilmez/daraltılmaz."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    header = [str(c).strip() if c else "" for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    companies: list[Company] = []
    for row in rows[1:]:
        kod = row[idx.get("BIST_Kod", 0)]
        if not kod:
            continue
        unvan = row[idx.get("Sirket_Unvani", 1)]
        sehir = row[idx.get("Sehir", 2)] if "Sehir" in idx else None
        companies.append(Company(kod=kod, unvan=unvan or "", sehir=sehir))

    if len(companies) != 100:
        logger.warning(
            "BIST100 listesi 100 satır bekleniyordu, %d bulundu (%s). "
            "SABİT örneklem varsayımı bozulmuş olabilir.",
            len(companies),
            path,
        )
    return companies


def load_company_reference(path=config.SIRKETLER_REFERANS_XLSX) -> dict[str, Company]:
    """788 şirketlik referans listesini KOD -> Company sözlüğüne çevirir.

    Kaynak dosyada iki özel durum var:
      1. Alfabetik bölüm ayraçları (tek harf/karakterli satırlar, örn. "A", "1") — atlanır.
      2. Bazı satırlarda TEK şirket için BİRDEN FAZLA kod virgülle ayrılmış olarak
         tutuluyor (örn. "GARAN, TGB", "ISATR, ISBTR, ISCTR, ISKUR, TIB" — farklı pay
         sınıfları/eski kodlar). Her kod ayrı ayrı sözlüğe eklenir ki BIST100
         eşleştirmesi (örn. "GARAN") bu satırı bulabilsin.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))

    reference: dict[str, Company] = {}
    for row in rows:
        kod_hucresi, unvan = row[0], row[1]
        if not kod_hucresi or not unvan:
            continue
        unvan = str(unvan).strip()
        if len(unvan) < 4:  # başlık/ayraç satırlarını ele
            continue
        sehir = row[2] if len(row) > 2 else None
        denetci = row[3] if len(row) > 3 else None
        for kod in str(kod_hucresi).split(","):
            kod = kod.strip()
            if not kod:
                continue
            reference[kod.upper()] = Company(kod=kod, unvan=unvan, sehir=sehir, bagimsiz_denetim=denetci)

    return reference


def load_manual_seed(path=config.MANUAL_IR_URLS_JSON) -> dict[str, dict]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


# --------------------------------------------------------------------------
# Cache
# --------------------------------------------------------------------------

def load_cache(path=config.IR_URL_CACHE_FILE) -> dict[str, InvestorRelationsInfo]:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        raw = json.load(f)
    return {kod: InvestorRelationsInfo(**data) for kod, data in raw.items()}


def save_cache(cache: dict[str, InvestorRelationsInfo], path=config.IR_URL_CACHE_FILE) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    serializable = {kod: asdict(info) for kod, info in cache.items()}
    tmp_path = path.with_suffix(path.suffix + ".tmp")
    with tmp_path.open("w", encoding="utf-8") as f:
        json.dump(serializable, f, ensure_ascii=False, indent=2, sort_keys=True)
    tmp_path.replace(path)


# --------------------------------------------------------------------------
# Otomatik keşif (ağ erişimi gerektirir)
# --------------------------------------------------------------------------

_TURKISH_MAP = str.maketrans({"ı": "i", "İ": "I", "ğ": "g", "ü": "u", "ş": "s", "ö": "o", "ç": "c"})

_LEGAL_SUFFIXES = re.compile(
    r"\b(A\.?Ş\.?|T\.?A\.?Ş\.?|HOLDİNG|HOLDING|SANAYİ|SANAYI|TİCARET|TICARET|"
    r"VE|A\.?O\.?|GAYRİMENKUL|YATIRIM|ORTAKLIĞI|ORTAKLIGI)\b",
    re.IGNORECASE,
)


def _slugify(unvan: str) -> str:
    """Şirket ünvanından kabaca bir alan adı adayı türetir (örn. 'X SANAYİ A.Ş.' -> 'x')."""
    text = unvan.translate(_TURKISH_MAP)
    text = _LEGAL_SUFFIXES.sub(" ", text)
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    text = re.sub(r"[^a-zA-Z0-9]+", "", text)
    return text.lower()


_DOMAIN_TOKEN_STOPWORDS = {"turk", "turkiye", "genel", "as", "ao"}


def _company_domain_tokens(company: Company) -> list[str]:
    """Ünvandan, alan adında beklenebilecek ayırt edici (ASCII, küçük harf)
    kelimeleri türetir - arama sonucu doğrulaması için."""
    ascii_unvan = unicodedata.normalize(
        "NFKD", company.unvan.translate(_TURKISH_MAP)
    ).encode("ascii", "ignore").decode("ascii")
    tokens = [t.lower() for t in re.split(r"[^a-zA-Z0-9]+", ascii_unvan) if len(t) >= 4]
    return [t for t in tokens if t not in _DOMAIN_TOKEN_STOPWORDS]


def _url_matches_company(url: str, company: Company) -> bool:
    """Arama motorundan gelen bir URL'nin gerçekten bu şirkete ait olma
    olasılığı var mı diye kaba bir kontrol yapar.

    Arama motoru scraping'i (özellikle bot tespit edilince) tamamen alakasız
    "çöp" sonuçlar döndürebiliyor (örn. Bing, otomatik trafiği fark edince
    şirket sorgusuna Çince forum sayfaları döndürdü). Alan adının şirket
    ünvanıyla hiçbir ortak kelimesi yoksa sonuç güvenilmez sayılıp reddedilir.
    """
    netloc = urlparse(url).netloc.lower()
    if netloc.startswith("www."):
        netloc = netloc[4:]
    if not netloc:
        return False
    domain_core = netloc.split(".")[0]
    tokens = _company_domain_tokens(company) + [company.kod.lower()]
    for tok in tokens:
        probe = tok[:6] if len(tok) > 6 else tok
        if len(probe) >= 3 and (probe in domain_core or domain_core in tok):
            return True
    return False


def candidate_domains(company: Company) -> list[str]:
    """Kod ve ünvandan olası kurumsal alan adları üretir. Doğrulanmamış tahminlerdir."""
    kod_slug = company.kod.lower()
    unvan_slug = _slugify(company.unvan)
    first_word_slug = _slugify(company.unvan.split()[0]) if company.unvan else ""

    candidates = []
    for slug in dict.fromkeys([unvan_slug, first_word_slug, kod_slug]):  # sırayı korur, tekrarı eler
        if not slug:
            continue
        candidates.append(f"https://www.{slug}.com.tr")
        candidates.append(f"https://www.{slug}.com")
    return candidates


def probe_url(url: str, session: requests.Session, timeout: float = 8.0) -> bool:
    """URL'nin canlı olup olmadığını kontrol eder. Ağ kapalıysa False döner (istisna yutulur)."""
    try:
        resp = session.head(url, timeout=timeout, allow_redirects=True, headers={"User-Agent": config.HTTP_USER_AGENT})
        if resp.status_code >= 400:
            # Bazı sunucular HEAD'i reddeder; hafif bir GET ile tekrar dene.
            resp = session.get(url, timeout=timeout, allow_redirects=True, headers={"User-Agent": config.HTTP_USER_AGENT}, stream=True)
        return resp.status_code < 400
    except requests.RequestException as exc:
        logger.debug("probe_url(%s) başarısız: %s", url, exc)
        return False


# --------------------------------------------------------------------------
# Arama motoru fallback'i (alan adı tahmini başarısız olduğunda)
# --------------------------------------------------------------------------
# API anahtarı gerektirmeyen iki HTML arama arayüzü sırayla denenir (Bing
# önce): DuckDuckGo'nun html.duckduckgo.com arayüzü kendini tanıtan bot
# User-Agent'ını ("Scope3ResearchBot" - HTTP_USER_AGENT) engelliyor VE aynı
# oturumdan art arda birkaç istekten sonra hızlıca oran-sınırlamasına giriyor
# (bu ajanın 100 şirketlik taraması için tek başına güvenilir değil); bu
# yüzden iki motor da denenir ve biri boş dönerse diğerine geçilir. Her ikisi
# de başarısız olursa şirket "bulunamadı" olarak işaretlenir - akış durmaz.
_SEARCH_USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)

# Sonuçlarda çıkabilecek ama şirketin KENDİ sitesi olmayan haber/finans/sosyal
# medya portalları - arama fallback'inin şirket sitesi yerine bunları
# döndürmesini engellemek için elenir.
_SEARCH_RESULT_EXCLUDE_DOMAINS = {
    "kap.org.tr", "isyatirim.com.tr", "wikipedia.org", "linkedin.com",
    "twitter.com", "x.com", "facebook.com", "instagram.com", "youtube.com",
    "bloomberght.com", "investing.com", "borsagundem.com", "foreks.com",
    "tradingview.com", "duckduckgo.com", "bing.com", "bigpara.hurriyet.com.tr",
    "borsaistanbul.com", "mynet.com", "google.com", "sabah.com.tr",
    "ntv.com.tr", "haberturk.com", "sozcu.com.tr", "milliyet.com.tr",
}


def _bing_search(query: str, session: requests.Session, timeout: float = 10.0) -> list[str]:
    """Bing HTML sonuç sayfasından organik sonuç URL'lerini döner."""
    try:
        resp = session.get(
            "https://www.bing.com/search",
            params={"q": query},
            timeout=timeout,
            headers={"User-Agent": _SEARCH_USER_AGENT},
        )
        resp.raise_for_status()
    except requests.RequestException as exc:
        logger.debug("Bing araması başarısız (%s): %s", query, exc)
        return []

    if BeautifulSoup is None:
        return []

    results = []
    soup = BeautifulSoup(resp.text, "html.parser")
    for li in soup.select("li.b_algo"):
        a = li.find("a", href=True)
        if not a:
            continue
        href = a["href"]
        if "bing.com/ck/a" in href:
            # Bing bazı sonuçları /ck/a?...&u=a1<base64> şeklinde tıklama
            # takip bağlantısıyla sarmalıyor; gerçek URL'yi geri çözer.
            encoded = parse_qs(urlparse(href).query).get("u", [""])[0]
            if encoded.startswith("a1"):
                encoded = encoded[2:]
                try:
                    href = base64.urlsafe_b64decode(encoded + "=" * (-len(encoded) % 4)).decode("utf-8", "ignore")
                except (ValueError, UnicodeDecodeError):
                    continue
        results.append(href)
    return results


def _ddg_search(query: str, session: requests.Session, timeout: float = 10.0) -> list[str]:
    """DuckDuckGo HTML sonuç sayfasından organik sonuç URL'lerini döner."""
    try:
        resp = session.post(
            "https://html.duckduckgo.com/html/",
            data={"q": query},
            timeout=timeout,
            headers={"User-Agent": _SEARCH_USER_AGENT},
        )
        resp.raise_for_status()
    except requests.RequestException as exc:
        logger.debug("DuckDuckGo araması başarısız (%s): %s", query, exc)
        return []

    results = []
    for m in re.finditer(r'class="result__a"[^>]*href="([^"]+)"', resp.text):
        href = m.group(1)
        qs = parse_qs(urlparse(href).query)
        real_url = unquote(qs["uddg"][0]) if "uddg" in qs else href
        results.append(real_url)
    return results


_SEARCH_ENGINES = (_bing_search, _ddg_search)


def search_company_website(company: Company, session: requests.Session) -> str | None:
    """Alan adı tahmini (candidate_domains) başarısız olduğunda web araması ile
    şirketin resmi sitesini/yatırımcı ilişkileri sayfasını bulmaya çalışır.

    API anahtarı gerektirmeyen iki HTML arama motoru sırayla denenir. Haber/
    finans portalı gibi şirketin kendi sitesi olmayan sonuçlar elenir; kalan
    ilk sonuç döndürülür (otomatik tahmin gibi doğrulanmamış sayılır).
    """
    for query in (f"{company.unvan} yatırımcı ilişkileri", f"{company.unvan} kurumsal web sitesi"):
        for engine in _SEARCH_ENGINES:
            for url in engine(query, session):
                netloc = urlparse(url).netloc.lower()
                if netloc.startswith("www."):
                    netloc = netloc[4:]
                if not netloc:
                    continue
                if any(netloc == ex or netloc.endswith("." + ex) for ex in _SEARCH_RESULT_EXCLUDE_DOMAINS):
                    continue
                if not _url_matches_company(url, company):
                    logger.debug(
                        "Arama sonucu %s, %s (%s) ünvanıyla örtüşmüyor - reddedildi.",
                        url, company.kod, company.unvan,
                    )
                    continue
                return url
    return None


def discover_for_company(
    company: Company,
    session: requests.Session,
    manual_seed: dict[str, dict],
    search_fn: SearchFn | None = None,
) -> InvestorRelationsInfo:
    """Tek bir şirket için resmi site/IR URL'sini bulmaya çalışır.

    Öncelik sırası: manuel tohum > alan adı tahmini + canlılık kontrolü >
    (varsa) harici arama fonksiyonu > "bulunamadı".
    """
    if company.kod in manual_seed:
        seed = manual_seed[company.kod]
        return InvestorRelationsInfo(
            kod=company.kod,
            resmi_web_sitesi=seed.get("resmi_web_sitesi"),
            yatirimci_iliskileri_url=seed.get("yatirimci_iliskileri_url"),
            kaynak="manuel",
            dogrulandi=bool(seed.get("dogrulandi", False)),
            not_=seed.get("not"),
        )

    for url in candidate_domains(company):
        if probe_url(url, session):
            return InvestorRelationsInfo(
                kod=company.kod,
                resmi_web_sitesi=url,
                yatirimci_iliskileri_url=None,
                kaynak="otomatik-tahmin",
                dogrulandi=False,
                not_="Alan adı canlı bulundu; yatırımcı ilişkileri alt sayfası elle doğrulanmalı.",
            )

    if search_fn is not None:
        found = search_fn(company)
        if found:
            found_lower = found.lower()
            is_ir_url = any(kw in found_lower for kw in ("yatirimci", "yatırımcı", "investor"))
            return InvestorRelationsInfo(
                kod=company.kod,
                resmi_web_sitesi=None if is_ir_url else found,
                yatirimci_iliskileri_url=found if is_ir_url else None,
                kaynak="arama-motoru",
                dogrulandi=False,
                not_="Arama sonucu; doğrulanmalı.",
            )

    return InvestorRelationsInfo(
        kod=company.kod,
        resmi_web_sitesi=None,
        yatirimci_iliskileri_url=None,
        kaynak=None,
        dogrulandi=False,
        not_="Bulunamadı - manuel araştırma gerekir (ağ erişimi kapalıysa bu beklenen bir sonuçtur).",
    )


# --------------------------------------------------------------------------
# Parça bazlı çalıştırma
# --------------------------------------------------------------------------

def run(
    companies: list[Company],
    progress_store: ProgressStore,
    chunk_size: int = config.CHUNK_SIZE,
    search_fn: SearchFn | None = None,
) -> dict[str, InvestorRelationsInfo]:
    """Tüm şirketler için IR keşfini 15-20'lik parçalar halinde çalıştırır ve cache'ler.

    search_fn belirtilmezse (varsayılan), alan adı tahmini (candidate_domains)
    canlı bir site bulamadığında otomatik olarak `search_company_website` ile
    web araması fallback'i devreye girer - domain tahmininin tutmadığı
    şirketler için "bulunamadı" sonucunu azaltmak amacıyla.
    """
    cache = load_cache()
    manual_seed = load_manual_seed()
    session = requests.Session()
    if search_fn is None:
        search_fn = lambda company: search_company_website(company, session)  # noqa: E731

    def process(company: Company) -> None:
        if company.kod in cache and cache[company.kod].dogrulandi:
            return  # zaten doğrulanmış, tekrar sorgulama
        cache[company.kod] = discover_for_company(company, session, manual_seed, search_fn)

    def on_chunk_done(chunk_idx: int, total_chunks: int, progress) -> None:
        save_cache(cache)
        found = sum(1 for k in progress.completed if cache.get(k) and cache[k].resmi_web_sitesi)
        print(
            f"[inventory] Parça {chunk_idx}/{total_chunks} tamamlandı — "
            f"toplam işlenen: {len(progress.completed)}, site bulunan: {found}, "
            f"hata: {len(progress.errors)}"
        )

    run_in_chunks(
        items=companies,
        key_fn=lambda c: c.kod,
        process_fn=process,
        progress_store=progress_store,
        stage_name="inventory",
        chunk_size=chunk_size,
        on_chunk_done=on_chunk_done,
    )
    save_cache(cache)
    return cache


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    companies = load_bist100()
    reference = load_company_reference()

    missing_in_reference = [c.kod for c in companies if c.kod not in reference]
    if missing_in_reference:
        logger.warning(
            "BIST100 listesindeki %d kod, 788 şirketlik referans listesinde bulunamadı: %s",
            len(missing_in_reference),
            ", ".join(missing_in_reference),
        )

    progress_store = ProgressStore(config.PROGRESS_FILE)
    run(companies, progress_store)


if __name__ == "__main__":
    main()

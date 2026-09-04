"""Ağ erişimi olmadan çalışan uçtan uca duman testi (smoke test).

inventory.py / downloader.py'nin canlı ağ gerektiren kısımlarını atlar;
bunun yerine PyMuPDF ile sentetik PDF'ler üretip text_miner ->
maturity_classifier -> excel_writer zincirinin doğru çalıştığını doğrular.
Bu, downloader.py'nin ağ-bağımlı kısmı test edilemese bile geri kalan
mimarinin bu ortamda gerçekten çalıştığının kanıtıdır.
"""
from __future__ import annotations

import shutil
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import pymupdf

import config
from src.excel_writer import write_workbook
from src.maturity_classifier import classify_all
from src.models import Company, ReportRef, ReportStatus
from src.text_miner import mine_pdf

# Base14 (Helvetica) Türkçe karakterleri (ı, ğ, ş, ç, ö, ü) desteklemiyor —
# testte gerçekçi Türkçe metin kullanmak için Unicode kapsamlı bir TTF gömüyoruz.
_TEST_FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"


def _make_pdf(path: Path, pages: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    doc = pymupdf.open()
    for text in pages:
        page = doc.new_page()
        page.insert_font(fontname="F0", fontfile=_TEST_FONT_PATH)
        rect = pymupdf.Rect(72, 72, page.rect.width - 72, page.rect.height - 72)
        overflow = page.insert_textbox(rect, text, fontsize=11, fontname="F0")
        assert overflow >= 0, f"Test metni sayfaya sığmadı (taşma: {overflow}): {text[:50]}..."
    doc.save(path)
    doc.close()


def run_smoke_test() -> None:
    tmp_root = config.BASE_DIR / "data" / "downloads" / "_SMOKE_TEST"
    if tmp_root.exists():
        shutil.rmtree(tmp_root)

    companies = [
        Company(kod="TEST1", unvan="TEST BİR SANAYİ A.Ş.", sehir="İSTANBUL"),
        Company(kod="TEST2", unvan="TEST İKİ HOLDİNG A.Ş.", sehir="ANKARA"),
        Company(kod="TEST3", unvan="TEST ÜÇ GAYRİMENKUL A.Ş.", sehir="İZMİR"),
    ]

    # TEST1: zengin açıklama -> Metodolojili-Güvenceli beklenir (kategori + güvence + sayı)
    pdf1 = tmp_root / "TEST1.pdf"
    _make_pdf(pdf1, [
        "Giris sayfasi, ilgisiz icerik.",
        (
            "Kapsam 3 emisyonlarımız değer zinciri boyunca hesaplanmıştır. "
            "Kategori 1 (satın alınan mal ve hizmetler) ve Kategori 4 (yukarı yönlü taşıma) "
            "kapsamında toplam 12.345 tCO2e emisyon raporlanmıştır. Bu veriler bağımsız "
            "güvence denetiminden geçmiş olup ISO 14064-3 standardına uygundur."
        ),
    ])

    # TEST2: sadece nitel bahis -> Nitel beklenir
    pdf2 = tmp_root / "TEST2.pdf"
    _make_pdf(pdf2, [
        "Sürdürülebilirlik yaklaşımımız kapsamında Scope 3 emisyonlarını izlemeyi hedefliyoruz.",
    ])

    # TEST3: hiçbir Kapsam 3 bahsi yok -> Açıklama Yok beklenir
    pdf3 = tmp_root / "TEST3.pdf"
    _make_pdf(pdf3, ["Bu raporda sadece Kapsam 1 ve Kapsam 2 emisyonları yer almaktadır."])

    manifest = [
        ReportRef(kod="TEST1", yil=2024, rapor_turu=config.REPORT_TYPE_GONULLU, yerel_dosya=str(pdf1), durum=ReportStatus.INDIRILDI, kaynak_turu="TEST"),
        ReportRef(kod="TEST2", yil=2024, rapor_turu=config.REPORT_TYPE_GONULLU, yerel_dosya=str(pdf2), durum=ReportStatus.INDIRILDI, kaynak_turu="TEST"),
        ReportRef(kod="TEST3", yil=2024, rapor_turu=config.REPORT_TYPE_GONULLU, yerel_dosya=str(pdf3), durum=ReportStatus.INDIRILDI, kaynak_turu="TEST"),
        ReportRef(kod="TEST1", yil=2023, rapor_turu=config.REPORT_TYPE_FAALIYET, durum=ReportStatus.BULUNMADI),
    ]

    all_findings = []
    for ref in manifest:
        if ref.durum == ReportStatus.INDIRILDI:
            found = mine_pdf(Path(ref.yerel_dosya), ref.kod, ref.yil, ref.rapor_turu)
            print(f"  {ref.kod}: {len(found)} bulgu")
            all_findings.extend(found)

    assert all_findings, "Hiç bulgu üretilmedi — text_miner zinciri bozuk olabilir"

    maturity_results = classify_all(all_findings)
    by_kod = {m.kod: m for m in maturity_results}

    assert by_kod["TEST1"].etiket.value == "Metodolojili-Güvenceli", by_kod["TEST1"]
    assert 1 in by_kod["TEST1"].kategori_kapsami and 4 in by_kod["TEST1"].kategori_kapsami, by_kod["TEST1"]
    assert by_kod["TEST2"].etiket.value == "Nitel", by_kod["TEST2"]
    assert "TEST3" not in by_kod  # hiç bulgu yok -> classify_all'da grup oluşmaz

    out_path = config.OUTPUT_DIR / "_smoke_test_output.xlsx"
    write_workbook(out_path, companies, manifest, all_findings, maturity_results)
    assert out_path.exists() and out_path.stat().st_size > 0

    import openpyxl

    wb = openpyxl.load_workbook(out_path)
    assert wb.sheetnames == [
        "Ham Bulgular", "Şirket-Yıl-RaporTürü Özeti", "Kategori 1-15 Matrisi", "Eksik Raporlar",
    ], wb.sheetnames

    ozet = wb["Şirket-Yıl-RaporTürü Özeti"]
    beklenen_satir = len(companies) * len(config.SCAN_YEARS) * len(config.REPORT_TYPES) + 1
    assert ozet.max_row == beklenen_satir, (ozet.max_row, beklenen_satir)

    eksik = wb["Eksik Raporlar"]
    eksik_kodlar = [row[0].value for row in eksik.iter_rows(min_row=2)]
    assert "TEST1" in eksik_kodlar  # 2023 Faaliyet Raporu -> Rapor Yok
    assert "TEST3" in eksik_kodlar  # hiçbir yıl/tür için TEST3 manifest'te yok -> Rapor Yok olarak görünmeli

    shutil.rmtree(tmp_root)
    out_path.unlink()

    print("\nTÜM SMOKE TEST KONTROLLERİ GEÇTİ.")
    print(f"  TEST1 -> {by_kod['TEST1'].etiket.value} (kategoriler: {by_kod['TEST1'].kategori_kapsami})")
    print(f"  TEST2 -> {by_kod['TEST2'].etiket.value}")
    print("  TEST3 -> Açıklama Yok (bulgu yok, grup oluşmadı)")
    print(f"  Özet sekmesi satır sayısı doğru: {beklenen_satir}")


if __name__ == "__main__":
    run_smoke_test()

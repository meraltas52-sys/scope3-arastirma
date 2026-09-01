#!/usr/bin/env python3
"""TSRS/KGK rapor envanterini (150 rapor) SABİT BIST100 örneklemiyle karşılaştırır.

Bu, "İLK ADIM" olarak konuşulan tek seferlik bir kontrol betiğidir (pipeline'ın
sürekli çalışan bir parçası değildir):

  1. Envanterdeki hangi şirketler BIST100 DIŞI? (silinmez, sadece raporlanır)
  2. BIST100'deki hangi şirketlerin envanterde HİÇ raporu yok? (bir sonraki
     adımda downloader.py'nin şirket IR sitelerinden araması gereken liste)

Ağ erişimi GEREKTİRMEZ — sadece iki xlsx dosyasını karşılaştırır.

Kullanım:
    python tools/compare_tsrs_bist100.py
"""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import Workbook

import config
from src.downloader import diff_tsrs_with_bist100
from src.inventory import load_bist100, load_company_reference


def main() -> None:
    companies = load_bist100()
    reference = load_company_reference()
    companies_by_kod = {c.kod: c for c in companies}

    diff = diff_tsrs_with_bist100(companies_by_kod)
    disi = diff["bist100_disi_satirlar"]
    eksik = diff["bist100_raporu_olmayan_sirketler"]

    print(f"TSRS envanterinde BIST100 DIŞI şirketlere ait rapor sayısı: {len(disi)}")
    print(f"BIST100'de olup envanterde HİÇ raporu olmayan şirket sayısı: {len(eksik)}/{len(companies)}")
    print()
    print("Raporu eksik olan şirketler (sonraki adım: şirket IR sitelerinden aranacak):")
    for kod in eksik:
        unvan = companies_by_kod[kod].unvan if kod in companies_by_kod else reference.get(kod, "")
        print(f"  {kod:8s} {unvan}")

    out_path = config.OUTPUT_DIR / "tsrs_bist100_karsilastirma.xlsx"
    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("BIST100 Dışı Satırlar")
    ws1.append(["Kod", "Rapor Başlığı"])
    for row in disi:
        ws1.append([row["kod"], row["baslik"]])

    ws2 = wb.create_sheet("Raporu Eksik BIST100 Şirketleri")
    ws2.append(["Kod", "Şirket Ünvanı", "Şehir"])
    for kod in eksik:
        c = companies_by_kod.get(kod)
        ws2.append([kod, c.unvan if c else "", c.sehir if c else ""])

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    print(f"\nDetaylı çıktı yazıldı: {out_path}")


if __name__ == "__main__":
    main()
